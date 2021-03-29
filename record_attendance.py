import pickle
import os
import time
import sys
from train import Train
from recogizer import *
import openpyxl


class attendance_record():

    def __init__(self, faces_directory='Faces'):
        self.STUDENT_NAMES = []
        self.faces_directory = faces_directory

    def get_names(self, faces_directory):
        # get the names in the Faces directory
        for name in os.listdir(faces_directory):
            self.STUDENT_NAMES.append(name)
        return self.STUDENT_NAMES

    def create_file_if_not_exists(self):

        student_names = ['Student Name']

        # GET THE FILE NAME
        # get the program directory      ex:  c:\Users\user\Desktop\face_detect_file_program
        path = os.path.dirname(__file__)

        # split the string by /   ->  ['c:', 'Users', 'user', 'Desktop', 'face_detect_file_program']
        path = path.split('/')

        # get the last item which is the program file   -> 'face_detect_file_program'
        file_name = path[-1]
        # check if the excel file exists, if not, then generate a new one
        if os.path.exists(f'{file_name}//attendance.xlsx'):

            excel_file = openpyxl.load_workbook('attendance.xlsx')
            sheet = excel_file['Attendance']

            # get the names of students from "Faces directory"
            names = self.get_names(self.faces_directory)

            # if the number of students saved in Faces is different compared to
            # the list of names saved in the excel sheet
            # append the updated list of students to excel
            # if len(sheet['A']) - 1 < len(names):
            for name in names:
                student_names.append(name)
            for index, cell in enumerate(student_names, start=1):
                sheet.cell(row=index, column=1, value=cell)
            return excel_file
        # create an excel file
        excel_file = openpyxl.Workbook()

        # create a sheet
        sheet = excel_file.create_sheet(title='Attendance')

        # delete the Sheet named 'Sheet' which is created by default when creating the excel file
        sht = excel_file.get_sheet_by_name('Sheet')
        excel_file.remove_sheet(sht)

        # get the names of students from "Faces directory"
        names = self.get_names(self.faces_directory)
        # add each name is names to the student_names list
        for name in names:
            student_names.append(name)

        # add student_names to the sheet
        for index, cell in enumerate(student_names, start=1):
            sheet.cell(row=index, column=1, value=cell)

        return excel_file

    def take_attendance(self, recognized_names_list):
        # raise an exception if there are no student names saved in the 'Faces' directory
        # or if face_encodings dont exist
        names = os.listdir(self.faces_directory)

        # list to hold the Presence/Abscence
        # relevant to the order of names in the excel sheet
        student_presence = ["Present/Absent"]

        # create or load the excel file
        excel_file = self.create_file_if_not_exists()

        sheet = excel_file['Attendance']

        # list of student names from the sheet in excel file
        sheet_names = sheet['A']
        # ommit the heading 'Student Name'
        sheet_names = sheet_names[1:]

        # list to hold initial preseneces and abscences
        sheet_attendance = sheet['B']
        # ommit the heading 'Student Name'
        sheet_attendance = sheet_attendance[1:]

        sheet_attendance_list = []

        for cell in sheet_attendance:
            sheet_attendance_list.append(cell.value)

        # if student name from excel sheet is present in the
        # recognized names list, add 'Present' else add 'Absent'
        # to student_presence list
        for index, cell in enumerate(sheet_names):
            if cell.value in recognized_names_list or sheet_attendance_list[index] == 'Present':
                student_presence.append('Present')
            else:
                student_presence.append('Absent')
        # add student presence list as a column to the excel sheet
        for index, cell in enumerate(student_presence, start=1):
            sheet.cell(row=index, column=2, value=cell)

        # save the excel file
        excel_file.save('attendance.xlsx')




